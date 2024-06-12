import openpyxl
import numpy as np

# Excel dosyasının doğru yolunu buraya ekleyin.
# Aşağıda belirtiler excell dosyasının yoludur.Sistemden sisteme değişiklilik gösterebilir. 
excel_dosya_yolu = r'C:\Users\ilyqs\Downloads\soru1_2_data (1).xlsx'
wb = openpyxl.load_workbook(excel_dosya_yolu)
sheet = wb.active
data = np.array([[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)])  # min_row=2, başlık satırını atlamak için
min_pixel_value = data.min()
max_pixel_value = data.max()
L = 256
A_new = (data - min_pixel_value) / (max_pixel_value - min_pixel_value) * (L - 1)
wb_new = openpyxl.Workbook()
sheet_new = wb_new.active
for i, row in enumerate(A_new):
    for j, value in enumerate(row):
        sheet_new.cell(row=i+1, column=j+1, value=value)
# Sonuç dosyasının oluşacağı yolu buraya eklemeliyiz bu sayede yeni dosya istediğiniz yere  oluşturulacaktır.
sonuc_dosya_yolu = r'C:\Users\ilyqs\Downloads\sonuc.xlsx'
wb_new.save(sonuc_dosya_yolu)
print(f"Sonuç '{sonuc_dosya_yolu}' olarak kaydedildi.")
